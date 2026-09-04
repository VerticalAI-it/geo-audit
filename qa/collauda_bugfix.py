# -*- coding: utf-8 -*-
"""I dieci punti del documento di bug fix del 4 settembre 2026, verificati.

Un punto per volta, sui dati veri. Non scrive niente: apre le pagine e guarda
cosa contengono.
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import db
import server
import views
from fastapi.testclient import TestClient

fall = []
c = TestClient(server.app)

# un progetto vero, con criticità e pagine
prog = next((p for p in db._sb_progetti_tutti()
             if db._sb_audits_by_project(p["id"], limit=1, full=False)), None)
if not prog:
    raise SystemExit("nessun progetto con audit: non si può collaudare")
PID = prog["id"]
UTENTE = {"id": prog["user_id"], "email": "prova@verticalai.it", "app_metadata": {}}
server._current_user = lambda req: (UTENTE, None)
print(f"progetto di prova: {prog.get('domain')}\n")


def controlla(etichetta, condizione, dettaglio=""):
    print(f"   {'ok ' if condizione else 'NO '} {etichetta}")
    if not condizione:
        fall.append(f"{etichetta}{' — ' + dettaglio if dettaglio else ''}")


def pagina(url):
    r = c.get(url)
    if r.status_code != 200:
        fall.append(f"{url}: HTTP {r.status_code}")
        return ""
    return r.text


# ── BUG-01 · il tema chiaro ────────────────────────────────────────────────
print("BUG-01 · tema chiaro")
import io as _io
for nome in ("form", "gate", "waiting", "login", "auth_callback"):
    p = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                     "templates", f"{nome}.html")
    t = _io.open(p, encoding="utf-8").read()
    controlla(f"{nome}.html non impone il tema scuro", 'data-theme="dark"' not in t)
    if "radial-gradient(140% 100% at 70% -5%, #1E1A38" in t:
        controlla(f"{nome}.html ha la variante chiara del fondo",
                  '[data-theme="light"] body' in t)

# ── BUG-02 e BUG-03 · logout e passaggio al pannello ──────────────────────
print("\nBUG-02 · logout raggiungibile ovunque")
for nome, url in (("Dashboard", "/dashboard"), ("Pagina progetto", f"/project/{PID}")):
    controlla(f"{nome}: c'è «Esci»", "/auth/logout" in pagina(url))

print("\nBUG-03 · il link al pannello, solo a chi è del team")
controlla("un cliente NON lo vede", 'href="/admin"' not in pagina("/dashboard"))
server._current_user = lambda req: ({**UTENTE, "app_metadata": {"role": "admin"}}, None)
controlla("chi è del team lo vede", 'href="/admin"' in pagina("/dashboard"))
server._current_user = lambda req: (UTENTE, None)

# ── BUG-04 · il tetto delle pagine ────────────────────────────────────────
print("\nBUG-04 · quante pagine guarda il motore")
controlla("il tetto è una costante sola", server.MAX_PAGINE == 30,
          f"vale {server.MAX_PAGINE}")
import re as _re
sorgente = _io.open(os.path.join(os.path.dirname(os.path.dirname(
    os.path.abspath(__file__))), "server.py"), encoding="utf-8").read()
resti = _re.findall(r"run_audit,\s*\w+,\s*(\d+)", sorgente)
controlla("nessun numero scritto a mano è rimasto", not resti, f"trovati: {resti}")

# ── BUG-05 · Issue e Critici in Pages ─────────────────────────────────────
print("\nBUG-05 · i numeri di Pages portano alle criticità")
h = pagina(f"/project/{PID}?tab=pages")
controlla("le celle sono link a Opportunities", "conta-link" in h and "tab=opportunities" in h)
controlla("il filtro passa dall'indirizzo", "pagina=" in h)

# ── BUG-06 · i due KPI di Technical ───────────────────────────────────────
print("\nBUG-06 · i conteggi di Technical filtrano")
h = pagina(f"/project/{PID}?tab=technical")
controlla("i KPI sono attivabili", 'class="kpi kpi-filtro"' in h)
controlla("si usano anche da tastiera", 'tabindex="0"' in h)
controlla("le righe dichiarano il loro stato", 'data-stato="' in h)

# ── BUG-07a e 07b · Riepilogo ─────────────────────────────────────────────
print("\nBUG-07a · i numeri del riepilogo")
h = pagina(f"/project/{PID}?tab=audit")
controlla("«problemi» e «critici» sono link", h.count("conta-link") >= 1)

print("\nBUG-07b · «Mostra altro» sugli interventi")
azioni = (db._sb_audits_by_project(PID, limit=1, full=True) or [{}])[0].get("actions") or []
if len(azioni) > 7:
    controlla(f"con {len(azioni)} interventi compare il bottone", "mostra-altro" in h)
    controlla("gli interventi oltre il settimo ci sono, nascosti", 'class="interv-item oltre"' in h)
else:
    controlla(f"con {len(azioni)} interventi il bottone NON serve", "mostra-altro" not in h)

# ── BUG-08a · la riga che dice cosa fare ──────────────────────────────────
print("\nBUG-08a · le righe si aprono e spiegano")
h = pagina(f"/project/{PID}?tab=opportunities")
controlla("le righe sono apribili", "riga-issue" in h and "riga-rimedio" in h)
ultimo = db._sb_audits_by_project(PID, limit=1, full=True)
rimedi = views._rimedi_per_check(ultimo[0] if ultimo else None)
controlla("i testi «come si risolve» esistono", len(rimedi) > 0, f"{len(rimedi)} testi")
issues = [i for i in db._sb_issues_by_project(PID) if i.get("status") == "open"]
scoperte = [i for i in issues if not rimedi.get(i.get("check_id") or "")]
controlla(f"tutte le {len(issues)} criticità aperte hanno un rimedio",
          not scoperte, f"{len(scoperte)} senza")

# ── BUG-08b · l'export ────────────────────────────────────────────────────
print("\nBUG-08b · l'export in Excel, con dentro la soluzione")
controlla("il bottone c'è in Opportunities", "export.xlsx?cosa=criticita" in h)
controlla("e in Pages", "export.xlsx?cosa=pagine" in pagina(f"/project/{PID}?tab=pages"))

r = c.get(f"/project/{PID}/export.xlsx?cosa=criticita")
controlla("il file si scarica", r.status_code == 200, f"HTTP {r.status_code}")
if r.status_code == 200:
    controlla("è un vero xlsx (non un CSV rinominato)", r.content[:2] == b"PK")
    controlla("si chiama .xlsx", ".xlsx" in r.headers.get("content-disposition", ""))
    try:
        from openpyxl import load_workbook
        wb = load_workbook(_io.BytesIO(r.content))
        ws = wb.active
        intest = [c.value for c in ws[1]]
        controlla("porta la colonna «Come si risolve»", "Come si risolve" in intest)
        righe = list(ws.iter_rows(min_row=2, values_only=True))
        con_rimedio = [r_ for r_ in righe if r_[-1]]
        controlla(f"su {len(righe)} righe, {len(con_rimedio)} hanno il rimedio scritto",
                  len(righe) == 0 or len(con_rimedio) > 0)
    except ImportError:
        fall.append("openpyxl non installato: l'export non si può verificare")

# ⚠️ e il file di un progetto altrui non si scarica
print("\n   e la proprietà del progetto")
server._current_user = lambda req: ({"id": "un-altro", "email": "x@y.it",
                                     "app_metadata": {}}, None)
r = c.get(f"/project/{PID}/export.xlsx?cosa=criticita")
controlla("un altro utente riceve 404", r.status_code == 404, f"HTTP {r.status_code}")

print("\nESITO:", "TUTTO A POSTO" if not fall else f"PROBLEMI ({len(fall)})")
for f in fall:
    print("   -", f)
raise SystemExit(1 if fall else 0)
